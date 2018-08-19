import openpyxl
import os.path
from pypinyin import pinyin, lazy_pinyin, Style
path = input('変換するxlsxを入力してください')
if os.path.exists(path+'.xlsx'):
    book = openpyxl.load_workbook(path+'.xlsx')
    active_sheet = book.active
    for i in range(1,active_sheet.max_row + 1):
        kanji = active_sheet.cell(i,1).value
        pins = pinyin(kanji)
        sentence = ""
        for pin in pins:
            sentence += pin[0] + " "
        active_sheet.cell(i,2, value=sentence)
    book.save(path+'(converted).xlsx')
    print("変換完了")
else:
    print("ファイルが存在しません")

